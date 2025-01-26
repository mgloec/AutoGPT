from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.urls import reverse
from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from .models import Task, Team, Category
from .forms import TaskForm, CategoryForm
from django.core.exceptions import ValidationError

# Create your views here.

@login_required
def export_tasks_excel(request):
    # Check if user is a team manager
    managed_teams = Team.objects.filter(manager=request.user)
    if not managed_teams.exists():
        messages.error(request, 'Only team managers can export task lists.')
        return redirect('timetracker:task_list')
    
    # Create a new workbook
    wb = Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Define headers
    headers = [
        'Title', 'Description', 'Category', 'Status',
        'Start Time', 'End Time', 'Duration (minutes)', 'Owner'
    ]
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create a sheet for each team
    for team in managed_teams:
        # Get tasks for this team
        team_tasks = Task.objects.filter(team=team).order_by('-created_at')
        
        # Create a sheet for this team
        ws = wb.create_sheet(title=team.name[:31])  # Excel sheet names limited to 31 chars
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        for row, task in enumerate(team_tasks, 2):
            ws.cell(row=row, column=1).value = task.title
            ws.cell(row=row, column=2).value = task.description
            ws.cell(row=row, column=3).value = task.category.name
            ws.cell(row=row, column=4).value = dict(Task.STATUS_CHOICES)[task.status]
            ws.cell(row=row, column=5).value = task.start_time.strftime('%Y-%m-%d %H:%M:%S') if task.start_time else ''
            ws.cell(row=row, column=6).value = task.end_time.strftime('%Y-%m-%d %H:%M:%S') if task.end_time else ''
            ws.cell(row=row, column=7).value = task.duration()
            ws.cell(row=row, column=8).value = task.owner.get_full_name() or task.owner.username
        
        # Add total row
        total_row = len(team_tasks) + 2
        ws.cell(row=total_row, column=1).value = "Total"
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        total_duration = sum((task.duration() for task in team_tasks), 0)
        ws.cell(row=total_row, column=7).value = total_duration
        ws.cell(row=total_row, column=7).font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = min(adjusted_width, 50)  # Cap width at 50
    
    # Create summary sheet
    summary = wb.create_sheet(title='Summary', index=0)
    summary.cell(row=1, column=1).value = "Team Summary"
    summary.cell(row=1, column=1).font = Font(bold=True)
    
    summary_headers = ['Team', 'Total Tasks', 'Total Duration (minutes)']
    for col, header in enumerate(summary_headers, 1):
        cell = summary.cell(row=2, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for row, team in enumerate(managed_teams, 3):
        team_tasks = Task.objects.filter(team=team)
        total_duration = sum((task.duration() for task in team_tasks), 0)
        
        summary.cell(row=row, column=1).value = team.name
        summary.cell(row=row, column=2).value = team_tasks.count()
        summary.cell(row=row, column=3).value = total_duration
    
    # Auto-adjust summary column widths
    for column in summary.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        summary.column_dimensions[column[0].column_letter].width = min(adjusted_width, 50)
    
    # Create the HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=tasks_report.xlsx'
    
    wb.save(response)
    return response

@login_required
def task_list(request):
    # Get teams where user is either a member or manager
    managed_teams = Team.objects.filter(manager=request.user)
    member_teams = Team.objects.filter(members=request.user)
    
    # Get the selected team from query parameters
    selected_team_id = request.GET.get('team')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # If user is a manager, show all tasks from their teams
    # If user is a member, only show their own tasks
    if managed_teams.exists():
        task_list = Task.objects.filter(team__in=managed_teams)
        is_manager = True
        available_teams = managed_teams
    else:
        task_list = Task.objects.filter(
            team__in=member_teams,
            owner=request.user
        )
        is_manager = False
        available_teams = member_teams
    
    # Apply team filter if selected
    if selected_team_id:
        try:
            selected_team = available_teams.get(id=selected_team_id)
            task_list = task_list.filter(team=selected_team)
        except Team.DoesNotExist:
            messages.error(request, 'Invalid team selected.')
    
    # Apply date filters if provided
    if start_date:
        task_list = task_list.filter(start_time__date__gte=start_date)
    if end_date:
        task_list = task_list.filter(start_time__date__lte=end_date)
    
    # Calculate total duration for all tasks in the filtered list
    total_duration = sum((task.duration() for task in task_list), 0)  # Sum in minutes
    total_seconds = total_duration * 60  # Convert to seconds for consistent display
    
    # Order tasks
    task_list = task_list.order_by('-created_at')
    
    # Number of tasks per page
    tasks_per_page = 10
    
    paginator = Paginator(task_list, tasks_per_page)
    page = request.GET.get('page')
    
    try:
        tasks = paginator.page(page)
    except PageNotAnInteger:
        tasks = paginator.page(1)
    except EmptyPage:
        tasks = paginator.page(paginator.num_pages)
    
    return render(request, 'timetracker/task_list.html', {
        'tasks': tasks,
        'is_manager': is_manager,
        'managed_teams': managed_teams,
        'available_teams': available_teams,
        'selected_team_id': selected_team_id,
        'total_duration_seconds': total_seconds,
        'start_date': start_date,
        'end_date': end_date,
    })

@login_required
def select_team(request):
    # Show all teams where user is either a member or manager
    teams = Team.objects.filter(
        Q(members=request.user) | Q(manager=request.user)
    ).distinct()
    return render(request, 'timetracker/select_team.html', {'teams': teams})

@login_required
def select_team_categories(request):
    # Get teams where user is manager
    teams = Team.objects.filter(manager=request.user).order_by('name')
    
    return render(request, 'timetracker/select_team_categories.html', {
        'teams': teams,
    })

@login_required
def task_create(request, team_id):
    team = get_object_or_404(Team, id=team_id)
    
    # Check if user is a member or manager of the team
    if not (team.members.filter(id=request.user.id).exists() or team.manager == request.user):
        messages.error(request, 'You must be a member or manager of this team to create tasks.')
        return redirect('timetracker:task_list')
    
    if request.method == 'POST':
        form = TaskForm(request.POST, team=team)
        if form.is_valid():
            task = form.save(commit=False)
            task.owner = request.user
            task.team = team
            
            # Set initial status as not_started
            task.status = 'not_started'
            # If start_time is set but no end_time, mark as in_progress
            if task.start_time and not task.end_time:
                task.status = 'in_progress'
            # If both start_time and end_time are set, mark as completed
            elif task.start_time and task.end_time:
                task.status = 'completed'
            
            task.save()
            messages.success(request, 'Task created successfully!')
            return redirect('timetracker:task_list')
        else:
            # Add form-level errors to messages if any exist
            for error in form.non_field_errors():
                messages.error(request, error)
    else:
        form = TaskForm(team=team)
    
    return render(request, 'timetracker/task_form.html', {
        'form': form, 
        'action': 'Create',
        'team': team
    })

@login_required
def task_edit(request, pk):
    task = get_object_or_404(Task, pk=pk)
    
    # Check if user is the owner or the team manager
    if not (task.owner == request.user or task.team.manager == request.user):
        messages.error(request, 'You must be the task owner or team manager to edit this task.')
        return redirect('timetracker:task_list')
    
    if request.method == 'POST':
        form = TaskForm(request.POST, instance=task, team=task.team)
        if form.is_valid():
            task = form.save(commit=False)
            task.status = form.cleaned_data.get('status', task.status)
            task.save()
            messages.success(request, 'Task updated successfully!')
            return redirect('timetracker:task_list')
    else:
        form = TaskForm(instance=task, team=task.team)
    
    return render(request, 'timetracker/task_form.html', {
        'form': form,
        'action': 'Edit',
        'team': task.team
    })

@login_required
def task_start(request, pk):
    task = get_object_or_404(Task, pk=pk)
    
    # Check if user is the owner
    if task.owner != request.user:
        return JsonResponse({'error': 'Only the task owner can start this task.'}, status=403)
    
    if task.start_time:
        return JsonResponse({'error': 'Task already started'}, status=400)
    
    task.start_time = timezone.now()
    task.status = 'in_progress'
    task.save()
    
    return JsonResponse({
        'start_time': task.start_time.isoformat(),
        'status': task.status,
        'duration': task.duration()
    })

@login_required
def task_stop(request, pk):
    task = get_object_or_404(Task, pk=pk)
    
    # Check if user is the owner
    if task.owner != request.user:
        return JsonResponse({'error': 'Only the task owner can stop this task.'}, status=403)
    
    if not task.start_time:
        return JsonResponse({'error': 'Task not started yet'}, status=400)
    
    if task.end_time:
        return JsonResponse({'error': 'Task already completed'}, status=400)
    
    task.end_time = timezone.now()
    task.status = 'completed'
    task.save()
    
    return JsonResponse({
        'end_time': task.end_time.isoformat(),
        'status': task.status,
        'duration': task.duration()
    })

@login_required
def task_delete(request, pk):
    task = get_object_or_404(Task, pk=pk)
    
    # Check if user is the owner or team manager
    if not (task.owner == request.user or task.team.manager == request.user):
        messages.error(request, 'You must be the task owner or team manager to delete this task.')
        return redirect('timetracker:task_list')
    
    task.delete()
    messages.success(request, 'Task deleted successfully!')
    return redirect('timetracker:task_list')

@login_required
def category_manage(request, team_id):
    team = get_object_or_404(Team, pk=team_id)
    
    # Check if user is the team manager
    if team.manager != request.user:
        messages.error(request, 'Only team managers can manage categories.')
        return redirect('timetracker:task_list')
    
    categories = Category.objects.filter(team=team).order_by('name')
    form = CategoryForm()
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add':
            form = CategoryForm(request.POST)
            if form.is_valid():
                category = form.save(commit=False)
                category.team = team
                category.save()
                messages.success(request, 'Category added successfully.')
                return redirect('timetracker:category_manage', team_id=team_id)
        
        elif action == 'edit':
            category_id = request.POST.get('category_id')
            category = get_object_or_404(Category, pk=category_id, team=team)
            category.name = request.POST.get('name')
            category.description = request.POST.get('description')
            category.save()
            messages.success(request, 'Category updated successfully.')
            return redirect('timetracker:category_manage', team_id=team_id)
        
        elif action == 'delete':
            category_id = request.POST.get('category_id')
            category = get_object_or_404(Category, pk=category_id, team=team)
            # Check if category is being used by any tasks
            if Task.objects.filter(category=category).exists():
                messages.warning(request, 'Cannot delete category as it is being used by tasks.')
            else:
                category.delete()
                messages.success(request, 'Category deleted successfully.')
            return redirect('timetracker:category_manage', team_id=team_id)
    
    return render(request, 'timetracker/category_manage.html', {
        'team': team,
        'categories': categories,
        'form': form,
    })
